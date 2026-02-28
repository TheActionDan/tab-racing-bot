"""
TAB Racing Bot
==============
Fetches live race data from TAB's API, enriches it with Racenet/PuntAPI form data,
sends everything to Claude AI for analysis, and outputs a formatted Excel spreadsheet
with picks for every race.

New in this version:
  - Track condition / going: WET TRACKER / DRY PREFERRED identification
  - Jockey and trainer form stats (via PuntAPI Phase 3 — experimental)
  - Horse distance step-up / step-down detection
  - Barrier success rate from PuntAPI (horse's record from today's draw)
  - Grade levelling — DROPS/RISES IN CLASS detection
  - Racing.com integration: speed ratings, track/distance/class/jockey stats, weight change
  - Horse weight from TAB API
  - Email automation via Gmail SMTP
  - Windows Task Scheduler setup (--setup flag)
  - API key moved to environment variable (ANTHROPIC_API_KEY)
  - Model switched from Opus to Sonnet (faster, ~80% cheaper)
  - Token-optimised prompts — target <20k tokens per batch

Requirements:
    pip install requests openpyxl anthropic

Environment variables (set before running):
    ANTHROPIC_API_KEY    — Claude API key (required)
    RACING_COM_API_KEY   — Racing.com GraphQL API key (required)
    GMAIL_USER           — Gmail address to send from (optional)
    GMAIL_APP_PASSWORD   — Gmail App Password, not your login (optional)
    EMAIL_RECIPIENTS     — Comma-separated recipient addresses (optional)

Windows quick-setup:
    setx ANTHROPIC_API_KEY "sk-ant-api03-..."
    setx RACING_COM_API_KEY "your-racing-com-api-key"
    setx GMAIL_USER "yourname@gmail.com"
    setx GMAIL_APP_PASSWORD "xxxx xxxx xxxx xxxx"
    setx EMAIL_RECIPIENTS "you@example.com,friend@example.com"
    (restart terminal after setx)

Usage:
    python tab_racing_bot.py
    python tab_racing_bot.py --date 2026-02-28
    python tab_racing_bot.py --date 2026-02-28 --state VIC
    python tab_racing_bot.py --setup   # create Task Scheduler batch file
"""

import sys
import os
import json
import argparse
import subprocess
import tempfile
import smtplib
import requests
import anthropic
from datetime import date, datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Force UTF-8 output so emoji/arrows print correctly on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


# ============================================================
# CONFIGURATION
# ============================================================

# Claude API key — load from environment variable.
# Windows: setx ANTHROPIC_API_KEY "sk-ant-api03-..."   then restart terminal.
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
if not ANTHROPIC_API_KEY:
    print("ERROR: ANTHROPIC_API_KEY environment variable is not set.")
    print("  Windows: setx ANTHROPIC_API_KEY \"sk-ant-api03-...\"")
    print("  Then close and reopen your terminal, and try again.")
    sys.exit(1)

# Model — Sonnet is ~5x faster and ~80% cheaper than Opus for this task
CLAUDE_MODEL = "claude-sonnet-4-6"

# TAB API base URL
TAB_API_BASE = "https://api.beta.tab.com.au/v1/tab-info-service/racing"

# PuntAPI (Racenet) GraphQL endpoint
PUNTAPI_BASE = "https://puntapi.com/racing"

# Racing.com GraphQL API — sectional times, speed ratings, per-stat breakdowns
RACING_COM_BASE    = "https://graphql.rmdprod.racing.com/"
RACING_COM_API_KEY = os.environ.get("RACING_COM_API_KEY", "")

# TAB headers — mimic a real browser
TAB_HEADERS = {
    "accept": "application/json, text/plain, */*",
    "accept-language": "en-GB,en-US;q=0.9,en;q=0.8",
    "origin": "https://www.tab.com.au",
    "referer": "https://www.tab.com.au/",
    "user-agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36"
    ),
    "sec-ch-ua": '"Not:A-Brand";v="99", "Google Chrome";v="145", "Chromium";v="145"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"Windows"',
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-site",
}

# Email config from environment variables (all optional)
GMAIL_USER       = os.environ.get("GMAIL_USER", "")
GMAIL_APP_PASS   = os.environ.get("GMAIL_APP_PASSWORD", "")
EMAIL_RECIPIENTS = [
    e.strip()
    for e in os.environ.get("EMAIL_RECIPIENTS", "").split(",")
    if e.strip()
]

# Races per Claude API call — 20 keeps comfortably under 20k token target
BATCH_SIZE = 20

# ── Track whitelist ────────────────────────────────────────────────────────────
# Location codes from the TAB API that we want to analyse.
# Override with --all-tracks to include all international meetings.
#   AU state codes: NSW VIC QLD SA WA NT TAS ACT
#   New Zealand:    NZL
#   Japan:          JPN
ALLOWED_LOCATIONS = {
    # Australian states / territories
    "NSW", "VIC", "QLD", "SA", "WA", "TAS", "ACT", "NT",
    # New Zealand
    "NZL",
    # Japan
    "JPN",
}

# ── Track barrier bias ─────────────────────────────────────────────────────────
# Known bias tendencies for major Australian tracks.
# Format: (bias_type, good_barrier_max, description)
#   "inside"  = inside barriers have a clear statistical advantage
#   "even"    = no strong draw bias
#   good_barrier_max = barriers up to this number are considered favourable
TRACK_BARRIER_BIAS = {
    "MOONEE VALLEY":  ("inside", 4,  "Very tight 1-turn circuit — inside draws critical"),
    "DOOMBEN":        ("inside", 4,  "Tight 1-turn track — inside draws dominant"),
    "EAGLE FARM":     ("inside", 5,  "Tight track — inside draws favoured"),
    "CAULFIELD":      ("inside", 5,  "Sharp turns — inside/middle preferred"),
    "MORPHETTVILLE":  ("inside", 5,  "Right-hand track — inside draws preferred"),
    "ROSEHILL":       ("inside", 6,  "Right-hand track — moderate inside bias"),
    "RANDWICK":       ("even",   7,  "Wide track — moderate inside advantage"),
    "ASCOT":          ("even",   7,  "Long track — moderate inside advantage"),
    "FLEMINGTON":     ("even",   8,  "Wide straight track — low draw bias"),
}


# ============================================================
# TRACK CONDITION HELPERS
# ============================================================

_WET_KEYWORDS = ("soft", "heavy")


def is_wet_track(condition: str) -> bool:
    """Return True if the track is Soft or Heavy (wet/rain-affected)."""
    if not condition:
        return False
    c = condition.lower()
    return any(w in c for w in _WET_KEYWORDS)


def compute_wet_preference(dry, wet) -> str:
    """
    Derive a surface preference label from PuntAPI place splits.
    dry / wet are each [1sts, 2nds, 3rds] on that surface.

    Returns 'WET TRACKER', 'DRY PREFERRED', or ''.
    Uses top-3 place finishes as a proxy for run count since we don't
    have total runs broken down by surface.
    """
    if not isinstance(dry, list):
        dry = [0, 0, 0]
    if not isinstance(wet, list):
        wet = [0, 0, 0]

    dry_tot = sum(dry[:3])
    wet_tot = sum(wet[:3])
    dry_w   = dry[0] if dry else 0
    wet_w   = wet[0] if wet else 0

    # Wet tracker: wins at least 25% on wet, and at least as good as dry
    if wet_tot >= 2 and wet_w >= 1:
        wet_rate = wet_w / wet_tot
        dry_rate = dry_w / dry_tot if dry_tot >= 2 else 0.0
        if wet_rate >= 0.25 and wet_rate >= dry_rate:
            return "WET TRACKER"

    # Dry preferred: clear dry record, little or no wet form
    if dry_tot >= 3 and dry_w >= 1:
        dry_rate = dry_w / dry_tot
        wet_rate = wet_w / wet_tot if wet_tot >= 1 else 0.0
        if dry_rate > wet_rate or (wet_tot == 0 and dry_w >= 1):
            return "DRY PREFERRED"

    return ""


def _class_difficulty_score(class_str: str):
    """
    Convert a race class string to a numeric difficulty score.
    HIGHER score = HARDER / more prestigious race.

    Scale (hardest → easiest):
      Group 1=520, Group 2=510, Group 3=500
      Listed=400
      Class 1=340, Class 2=330, Class 3=320, Class 4=310, Class 5=300
      BM xx = 100 + actual_BM_number  (BM64=164, BM80=180, BM100=200, etc.)
      Maiden=1

    NOTE: Class numbers are INVERTED in Australian racing —
          Class 1 is the HARDEST, Class 5 the easiest.
          BenchMark numbers are NORMAL — higher BM number = harder race.
    """
    import re as _re
    if not class_str:
        return None
    s = str(class_str).strip().upper()

    # Maiden — lowest grade
    if 'MAIDEN' in s:
        return 1

    # BenchMark / BM races: higher number = harder
    # Handles both "BM84" and "BenchMark 84" (API uses full word)
    m = _re.search(r'(?:BENCHMARK|BM)\s*(\d+)', s)
    if m:
        return 100 + int(m.group(1))

    # "0 - XX" rating bands (e.g. "0 - 58", "0 - 64") — treat like BM of that rating
    m = _re.match(r'^0\s*[-\u2013]\s*(\d+)$', s)
    if m:
        return 100 + int(m.group(1))

    # Class races: Class 1 = hardest (340), Class 5 = easiest (300)
    # Class 6 and beyond continue the scale downward (290, 280, ...)
    m = _re.search(r'\bCLASS\s*(\d+)\b', s)
    if m:
        n = int(m.group(1))
        return 300 + (5 - n) * 10  # Class 1=340, Class 2=330, ..., Class 5=300, Class 6=290

    # Listed: between Class 1 (340) and Group 3 (500)
    if 'LISTED' in s:
        return 400

    # Group races: Group 1 = hardest (520), Group 3 = easiest (500)
    m = _re.search(r'\bGR(?:OUP|P?)\.?\s*(\d)\b', s)
    if m:
        n = int(m.group(1))
        return 500 + (3 - n) * 10  # Group 1=520, Group 2=510, Group 3=500

    return None


def compute_grade_change(current_class, last_class) -> str:
    """
    Compare today's race class against last run class.
    Returns 'DROPS IN CLASS', 'RISES IN CLASS', 'SAME CLASS', or ''.

    Australian class hierarchy (hardest → easiest):
      Group 1 > Group 2 > Group 3 > Listed >
      Class 1 > Class 2 > Class 3 > Class 4 > Class 5 >
      BM 100+ > BM 90 > BM 80 > BM 70 > BM 64 > Maiden

    Class numbers are INVERTED: Class 1 is the hardest, Class 5 the easiest.
    BenchMark numbers are normal: higher BM = harder race.
    """
    if not current_class or not last_class:
        return ""
    curr = str(current_class).strip()
    last = str(last_class).strip()
    if curr.upper() == last.upper():
        return "SAME CLASS"
    c_score = _class_difficulty_score(curr)
    l_score = _class_difficulty_score(last)
    if c_score is not None and l_score is not None:
        if c_score < l_score:
            return f"DROPS IN CLASS ({last} -> {curr})"
        elif c_score > l_score:
            return f"RISES IN CLASS ({last} -> {curr})"
        else:
            return "SAME CLASS"
    # Fallback: show the change without a direction call
    return f"CLASS: {last} -> {curr}"


# ── PuntAPI helper utilities ──────────────────────────────────────────────────

def _parse_stats_class(raw: str):
    """
    PuntAPI stats.class returns a string like '3:0-1-1'
    (class_level:wins-seconds-thirds). Extract just the class level prefix.
    Returns e.g. 'Class 3' or None if unparseable.
    """
    if not raw:
        return None
    try:
        level = raw.split(":")[0].strip()
        return f"Class {level}" if level.isdigit() else raw
    except Exception:
        return raw


def _index_barrier_stats(raw) -> dict:
    """
    Convert barrierStats list [{name, wins, runs}, ...] into a dict keyed by
    barrier name (string) for O(1) lookup by today's barrier number.
    Handles None / empty gracefully.
    """
    if not raw or not isinstance(raw, list):
        return {}
    result = {}
    for entry in raw:
        if isinstance(entry, dict) and entry.get("name") is not None:
            result[str(entry["name"])] = entry
    return result


# ============================================================
# RACING.COM API HELPERS
# ============================================================

_RACING_COM_MEETINGS_QUERY = """{
  GetMeetingByDate(date: "%date%") {
    id
    venueName
    state
  }
}
"""

_RACING_COM_RACES_QUERY = """{
  getRacesForMeet(meetCode: "%meet_code%") {
    raceNumber
    formRaceEntries {
      horseName
      barrierNumber
      speedValue
      atThisBarrierNumberStats
      atThisClassStats
      jockeyStats
      trackStats
      distanceStats
      weightCarried
      weightPrevious
    }
  }
}
"""


def _racing_com_gql(query: str) -> dict:
    """
    Execute a GraphQL query against the Racing.com API.
    Uses curl with a temp file to avoid shell-escaping issues on Windows.
    """
    payload = json.dumps({"query": query})
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", delete=False, encoding="utf-8"
    )
    try:
        tmp.write(payload)
        tmp.close()
        result = subprocess.run(
            [
                "curl", "-s", "-X", "POST", RACING_COM_BASE,
                "-H", "Content-Type: application/json",
                "-H", f"x-api-key: {RACING_COM_API_KEY}",
                "-d", f"@{tmp.name}",
            ],
            capture_output=True, text=True, timeout=60,
        )
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass

    stdout = result.stdout.strip()
    if not stdout:
        return {}
    try:
        return json.loads(stdout)
    except json.JSONDecodeError:
        return {}


def _parse_rc_stat(stat_str) -> dict:
    """
    Parse a Racing.com stat string "runs:wins-2nds-3rds" into a dict.
    Returns {runs, wins, seconds, thirds} or {} if unparseable.
    Example: "3:1-1-0" → {runs:3, wins:1, seconds:1, thirds:0}
    """
    if not stat_str:
        return {}
    try:
        colon = stat_str.index(":")
        runs   = int(stat_str[:colon])
        parts  = stat_str[colon + 1:].split("-")
        wins   = int(parts[0]) if parts else 0
        secs   = int(parts[1]) if len(parts) > 1 else 0
        thirds = int(parts[2]) if len(parts) > 2 else 0
        return {"runs": runs, "wins": wins, "seconds": secs, "thirds": thirds}
    except (ValueError, AttributeError, IndexError):
        return {}


def _fmt_rc_stat(stat_str) -> str:
    """Format a Racing.com stat string for display. '3:1-1-0' → '1W/3R(33%)'."""
    s = _parse_rc_stat(stat_str)
    if not s or s.get("runs", 0) == 0:
        return ""
    r, w = s["runs"], s["wins"]
    pct = w / r * 100
    return f"{w}W/{r}R({pct:.0f}%)"


def _compute_weight_change(current: str, previous: str) -> str:
    """
    Compare carried weight vs previous weight and return a label.
    e.g. 'Lighter 1.0kg', 'Heavier 2.5kg', or '' if no meaningful difference.
    """
    def parse_kg(s):
        if not s:
            return None
        try:
            return float(str(s).replace("kg", "").strip())
        except (ValueError, AttributeError):
            return None

    curr = parse_kg(current)
    prev = parse_kg(previous)
    if curr is None or prev is None:
        return ""
    diff = curr - prev
    if abs(diff) < 0.4:   # less than 400g — not significant
        return ""
    if diff < 0:
        return f"Lighter {abs(diff):.1f}kg"
    return f"Heavier {diff:.1f}kg"


def get_racing_com_form(race_date: str) -> dict:
    """
    Fetch Racing.com form data for all Australian meetings on race_date.

    Returns {HORSE_NAME_UPPER: {speed_rating, barrier_stats, class_stats,
                                 jockey_stats, track_stats, dist_stats,
                                 weight_today, weight_last}}
    """
    print("\nFetching Racing.com form data...")

    # Step 1: get all meetings for the date
    try:
        mtg_data = _racing_com_gql(
            _RACING_COM_MEETINGS_QUERY.replace("%date%", race_date)
        )
    except Exception as e:
        print(f"  Racing.com unavailable: {e}")
        return {}

    if not mtg_data or "errors" in mtg_data:
        err = (
            mtg_data.get("errors", [{}])[0].get("message", "unknown")
            if mtg_data else "no response"
        )
        print(f"  Racing.com error: {err}")
        return {}

    all_meetings = mtg_data.get("data", {}).get("GetMeetingByDate", [])
    aus_meetings = [m for m in all_meetings if m.get("state") in AUS_STATES]
    print(f"  Found {len(aus_meetings)} Australian meetings on Racing.com")

    # Step 2: for each meeting fetch all races and entries in one query
    racing_com_lookup = {}
    for meeting in aus_meetings:
        meet_id = meeting["id"]
        try:
            races_data = _racing_com_gql(
                _RACING_COM_RACES_QUERY.replace("%meet_code%", meet_id)
            )
        except Exception:
            continue
        if not races_data or "data" not in races_data:
            continue

        races = races_data.get("data", {}).get("getRacesForMeet", [])
        for race in races:
            for entry in (race.get("formRaceEntries") or []):
                horse_name = (entry.get("horseName") or "").strip()
                if not horse_name:
                    continue
                racing_com_lookup[horse_name.upper()] = {
                    "speed_rating":   entry.get("speedValue"),
                    "barrier_stats":  entry.get("atThisBarrierNumberStats"),
                    "class_stats":    entry.get("atThisClassStats"),
                    "jockey_stats":   entry.get("jockeyStats"),
                    "track_stats":    entry.get("trackStats"),
                    "dist_stats":     entry.get("distanceStats"),
                    "weight_today":   entry.get("weightCarried"),
                    "weight_last":    entry.get("weightPrevious"),
                }

    print(f"  Racing.com: {len(racing_com_lookup)} runners enriched")
    return racing_com_lookup


# ============================================================
# STEP 1: FETCH DATA FROM TAB API
# ============================================================

def get_meetings(race_date: str, jurisdiction: str = "NSW",
                 all_tracks: bool = False) -> list:
    """
    Fetch all race meetings for a given date and jurisdiction.
    'NSW' returns all national meetings (TAB's default national feed).

    By default only AU / NZ / Japan meetings are returned (ALLOWED_LOCATIONS).
    Pass all_tracks=True to override and include every country.
    """
    url    = f"{TAB_API_BASE}/dates/{race_date}/meetings"
    params = {"jurisdiction": jurisdiction, "returnOffers": "true", "returnPromo": "true"}
    print(f"Fetching meetings from TAB API for {race_date}...")
    response = requests.get(url, headers=TAB_HEADERS, params=params, timeout=15)
    response.raise_for_status()
    meetings       = response.json().get("meetings", [])
    horse_meetings = [m for m in meetings if m.get("raceType") == "R"]

    if all_tracks:
        print(
            f"Found {len(horse_meetings)} horse racing meetings "
            "(all countries — whitelist disabled)"
        )
        return horse_meetings

    # Apply whitelist: keep only AU + NZ + JPN
    filtered   = [m for m in horse_meetings if m.get("location") in ALLOWED_LOCATIONS]
    dropped    = [m.get("meetingName", "?") for m in horse_meetings
                  if m.get("location") not in ALLOWED_LOCATIONS]
    print(
        f"Found {len(filtered)} meetings (AU/NZ/JPN) from {len(horse_meetings)} total "
        f"— skipped {len(dropped)} international: {', '.join(dropped) or 'none'}"
    )
    return filtered


def get_race_detail(venue_mnemonic: str, race_type: str, race_number: int,
                    race_date: str, jurisdiction: str = "NSW") -> dict:
    """Fetch detailed runner/odds data for a specific race."""
    url = (
        f"{TAB_API_BASE}/dates/{race_date}/meetings"
        f"/{race_type}/{venue_mnemonic}/races/{race_number}"
    )
    response = requests.get(
        url, headers=TAB_HEADERS, params={"jurisdiction": jurisdiction}, timeout=15
    )
    return response.json() if response.status_code == 200 else {}


def _make_runner_skeleton(r: dict, fixed: dict, parimutuel: dict = None) -> dict:
    """Build the standard runner dict from raw TAB runner data."""
    parimutuel = parimutuel or {}
    return {
        "number":          r.get("runnerNumber", ""),
        "name":            r.get("runnerName", "Unknown"),
        "barrier":         r.get("barrierNumber", ""),
        "jockey":          r.get("riderDriverName", ""),
        "trainer":         r.get("trainerName", ""),
        # Weight — TAB may expose as handicapWeight or weightTotal (kg)
        "weight":          (
            r.get("handicapWeight")
            or r.get("weightTotal")
            or r.get("weightKg")
            or ""
        ),
        "win_fixed":       fixed.get("returnWin") or 0,
        "place_fixed":     fixed.get("returnPlace") or 0,
        "win_tote":        parimutuel.get("returnWin") or 0,
        "scratched":       False,
        # Form fields — filled by PuntAPI enrichment
        "career":          "",
        "dry":             "",
        "wet":             "",
        "last_runs":       "",
        "days_since":      None,
        # Track surface preference (computed from dry/wet splits)
        "wet_preference":  "",
        # Distance change vs last start
        "distance_change": "",
        # Jockey and trainer stats (from PuntAPI Phase 3 — may remain None)
        "jockey_wins":     None,
        "jockey_runs":     None,
        "trainer_wins":    None,
        "trainer_runs":    None,
        # Barrier stats — horse's record from today's specific barrier (PuntAPI experimental)
        "barrier_wins":       None,
        "barrier_runs":       None,
        "barrier_flag":       "",   # e.g. "BARRIER ADVANTAGE — 4W/8R (50%)"
        "track_barrier_note": "",   # e.g. "GOOD DRAW (3) — inside favoured"
        # Grade levelling
        "current_class":   None,
        "grade_change":    "",      # e.g. "DROPS IN CLASS (BM84 -> BM72)"
        # Racing.com enrichment — speed ratings and form breakdowns
        "speed_rating":    None,   # Speed figure / rank from Racing.com
        "rc_barrier_stats": None,  # "runs:W-P-S" at today's barrier (Racing.com)
        "rc_class_stats":  None,   # "runs:W-P-S" at this class level
        "rc_jockey_stats": None,   # "runs:W-P-S" jockey at this venue+distance
        "rc_track_stats":  None,   # "runs:W-P-S" at this track
        "rc_dist_stats":   None,   # "runs:W-P-S" at this distance
        "weight_change":   "",     # e.g. "Lighter 1.0kg" or "Heavier 2.0kg"
    }


def extract_runners(race_detail: dict) -> list:
    """Parse TAB race detail JSON and extract runner info."""
    runners = []
    for runner in race_detail.get("runners", []):
        fixed      = runner.get("fixedOdds", {})
        parimutuel = runner.get("parimutuel", {})
        status     = fixed.get("bettingStatus", "") or parimutuel.get("bettingStatus", "")
        if "Scratched" in status:
            continue
        runners.append(_make_runner_skeleton(runner, fixed, parimutuel))

    runners.sort(
        key=lambda x: (x["win_fixed"] or 0) if (x["win_fixed"] or 0) > 0
        else ((x["win_tote"] or 0) if (x["win_tote"] or 0) > 0 else 999)
    )
    return runners


def build_race_summary(
    meetings: list,
    race_date: str,
    jurisdiction: str,
    form_lookup: dict,
    jockey_lookup: dict,
    trainer_lookup: dict,
    racing_com_lookup: dict = None,
) -> list:
    """
    Loop through all meetings and races, fetch TAB runner data, then enrich
    each runner with PuntAPI form, track preference, jockey, and trainer data.
    """
    all_races = []
    for meeting in meetings:
        meeting_name    = meeting.get("meetingName", "")
        venue_mnemonic  = meeting.get("venueMnemonic", "")
        race_type       = meeting.get("raceType", "R")
        location        = meeting.get("location", "")
        track_condition = meeting.get("trackCondition", "")
        track_wet       = is_wet_track(track_condition)
        races           = meeting.get("races", [])

        print(
            f"\n  {meeting_name} ({location}) — {len(races)} races "
            f"| Track: {track_condition}{'  [WET TRACK]' if track_wet else ''}"
        )

        for race in races:
            race_number = race.get("raceNumber")
            race_name   = race.get("raceName", f"Race {race_number}")
            distance    = race.get("raceDistance") or 0
            start_time  = race.get("raceStartTime", "")

            detail  = get_race_detail(venue_mnemonic, race_type, race_number, race_date, jurisdiction)
            runners = extract_runners(detail)

            # Fallback: use meeting-level runner list if detail fetch returned nothing
            if not runners:
                for r in race.get("runners", []):
                    fixed  = r.get("fixedOdds", {})
                    status = fixed.get("bettingStatus", "")
                    if "Scratched" in status:
                        continue
                    runners.append(_make_runner_skeleton(r, fixed))
                runners.sort(key=lambda x: x["win_fixed"] if x["win_fixed"] > 0 else 999)

            # ── Enrich each runner ────────────────────────────────────────────
            for runner in runners:
                key  = runner["name"].upper()
                form = form_lookup.get(key, {})

                runner["career"]    = form.get("career", "")
                runner["dry"]       = form.get("dry", "")
                runner["wet"]       = form.get("wet", "")
                runner["last_runs"] = form.get("last_runs", "")
                runner["days_since"] = form.get("days_since")

                # Surface preference from raw array splits
                runner["wet_preference"] = compute_wet_preference(
                    form.get("dry_arr", []),
                    form.get("wet_arr", []),
                )

                # Distance step-up / step-down vs last start
                last_dist = form.get("last_distance")
                if last_dist and distance:
                    try:
                        diff = int(distance) - int(last_dist)
                        if abs(diff) >= 200:
                            direction = "UP" if diff > 0 else "DOWN"
                            runner["distance_change"] = f"Step {direction} {abs(diff)}m"
                        elif abs(diff) > 0:
                            runner["distance_change"] = f"Similar ({diff:+d}m)"
                    except (ValueError, TypeError):
                        pass

                # Jockey stats (Phase 3 — may be empty)
                jstats = jockey_lookup.get(runner["jockey"].upper(), {})
                if jstats.get("total_runs"):
                    runner["jockey_wins"] = jstats["wins"]
                    runner["jockey_runs"] = jstats["total_runs"]

                # Trainer stats (Phase 3 — may be empty)
                tstats = trainer_lookup.get(runner["trainer"].upper(), {})
                if tstats.get("total_runs"):
                    runner["trainer_wins"] = tstats["wins"]
                    runner["trainer_runs"] = tstats["total_runs"]

                # ── Barrier stats ────────────────────────────────────────────
                barrier_num = runner.get("barrier")
                bstats_raw  = form.get("barrier_stats_raw", {})

                # Match today's barrier number to the indexed dict
                bstats = bstats_raw.get(str(barrier_num), {})
                bw = bstats.get("wins")
                br = bstats.get("runs")
                runner["barrier_wins"] = bw
                runner["barrier_runs"] = br

                # Horse's own record from today's barrier
                if bw is not None and br and br >= 3:
                    pct = bw / br * 100
                    b_label = f"{bw}W/{br}R ({pct:.0f}%)"
                    if pct >= 40:
                        runner["barrier_flag"] = f"BARRIER ADVANTAGE — {b_label} from barrier {barrier_num}"
                    elif pct == 0:
                        runner["barrier_flag"] = f"BARRIER CONCERN — 0W/{br}R from barrier {barrier_num}"
                    else:
                        runner["barrier_flag"] = f"Barrier {barrier_num}: {b_label}"

                # Track barrier bias (hardcoded knowledge for known tracks)
                bias_type, good_max, bias_note = TRACK_BARRIER_BIAS.get(
                    meeting_name.upper(), ("even", 8, "")
                )
                try:
                    b = int(barrier_num)
                    if bias_type == "inside":
                        if b <= good_max:
                            runner["track_barrier_note"] = (
                                f"GOOD DRAW (B{b}) — {bias_note}"
                            )
                        else:
                            runner["track_barrier_note"] = (
                                f"WIDE DRAW (B{b}) — {bias_note}"
                            )
                except (ValueError, TypeError):
                    pass

                # ── Grade levelling ──────────────────────────────────────────
                runner["current_class"] = form.get("current_class")
                runner["grade_change"]  = compute_grade_change(
                    form.get("current_class"),
                    form.get("last_class"),
                )

                # ── Racing.com enrichment ─────────────────────────────────────
                if racing_com_lookup:
                    rc = racing_com_lookup.get(key, {})
                    runner["speed_rating"]    = rc.get("speed_rating")
                    runner["rc_barrier_stats"] = rc.get("barrier_stats")
                    runner["rc_class_stats"]  = rc.get("class_stats")
                    runner["rc_jockey_stats"] = rc.get("jockey_stats")
                    runner["rc_track_stats"]  = rc.get("track_stats")
                    runner["rc_dist_stats"]   = rc.get("dist_stats")
                    runner["weight_change"]   = _compute_weight_change(
                        rc.get("weight_today"), rc.get("weight_last")
                    )

            all_races.append({
                "track":           meeting_name,
                "location":        location,
                "track_condition": track_condition,
                "track_wet":       track_wet,
                "race_number":     race_number,
                "race_name":       race_name,
                "distance":        distance,
                "start_time":      start_time,
                "runners":         runners,
            })

    return all_races


# ============================================================
# STEP 1b: FETCH FORM DATA FROM PUNTAPI (RACENET)
# ============================================================

# Australian state codes used by PuntAPI
AUS_STATES = {"NSW", "QLD", "VIC", "SA", "WA", "TAS", "ACT", "NT"}

# Phase 1: bulk career stats across all meetings
_PUNTAPI_STATS_QUERY = """
{
  meetings(startDate: "%date%", endDate: "%date%") {
    id
    name
    state
    events {
      id
      eventNumber
      entryConditions { type description }
      selections {
        id
        barrierNumber
        weight
        status
        competitor { id name }
        jockey { id name }
        trainer { id name }
        stats {
          wins
          totalRuns
          dryPlaces
          wetPlaces
          class
          barrierStats { name wins runs }
        }
      }
    }
  }
}
"""

# Phase 2: per-meeting last-run details (adding in bulk would timeout)
_PUNTAPI_LASTRUN_QUERY = """
{
  meeting(id: "%meeting_id%") {
    id
    events {
      id
      selections {
        id
        competitor { id name }
        lastRun {
          id
          finishPosition
          margin
          meetingName
          event {
            name
            distance
            startTime
            entryConditions { type description }
          }
        }
      }
    }
  }
}
"""

# Phase 3 (experimental): jockey and trainer career stats.
# PuntAPI may or may not expose stats on jockey/trainer objects —
# this is tried and silently skipped if the schema doesn't support it.
_PUNTAPI_JOCKTRN_QUERY = """
{
  meetings(startDate: "%date%", endDate: "%date%") {
    id
    state
    events {
      id
      selections {
        id
        jockey {
          id
          name
          stats { wins totalRuns }
        }
        trainer {
          id
          name
          stats { wins totalRuns }
        }
      }
    }
  }
}
"""


def _puntapi_gql(query: str) -> dict:
    """
    Execute a GraphQL query against PuntAPI using curl.
    Writes payload to a temp file to avoid shell-escaping issues on Windows.
    """
    payload = json.dumps({"query": query})
    tmp = tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", delete=False, encoding="utf-8"
    )
    try:
        tmp.write(payload)
        tmp.close()
        result = subprocess.run(
            [
                "curl", "-s", "-X", "POST", PUNTAPI_BASE,
                "-H", (
                    "User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36"
                ),
                "-H", "Origin: https://www.racenet.com.au",
                "-H", "Content-Type: application/json",
                "-H", "Accept: application/json",
                "-H", "Authorization: Bearer guest",
                "-d", f"@{tmp.name}",
            ],
            capture_output=True, text=True, timeout=60,
        )
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass

    stdout = result.stdout.strip()
    if not stdout:
        return {}
    try:
        return json.loads(stdout)
    except json.JSONDecodeError:
        return {}


def get_puntapi_form(race_date: str):
    """
    Fetch Racenet form data for all meetings on race_date.

    Returns a tuple of three dicts:
      form_lookup    — {HORSE_NAME_UPPER: form_dict}
      jockey_lookup  — {JOCKEY_NAME_UPPER: {wins, total_runs}}  (may be {})
      trainer_lookup — {TRAINER_NAME_UPPER: {wins, total_runs}} (may be {})
    """
    print("\nFetching form data from PuntAPI (Racenet)...")
    today = datetime.fromisoformat(race_date)

    # ── Phase 1: career stats (bulk) ──────────────────────────────────────────
    try:
        data = _puntapi_gql(_PUNTAPI_STATS_QUERY.replace("%date%", race_date))
    except Exception as e:
        print(f"  PuntAPI unavailable: {e} — continuing without form data")
        return {}, {}, {}

    if not data or "errors" in data or "message" in data:
        msg = (
            data.get("errors", [{}])[0].get("message", data.get("message", "unknown"))
            if data else "no response"
        )
        print(f"  PuntAPI error: {msg} — continuing without form data")
        return {}, {}, {}

    meetings        = data.get("data", {}).get("meetings", [])
    form_lookup     = {}
    sel_by_id       = {}
    aus_meeting_ids = []

    for meeting in meetings:
        if meeting.get("state") in AUS_STATES:
            aus_meeting_ids.append(meeting["id"])
        for event in meeting.get("events", []):
            # Extract race class from entryConditions (type=="Class" entry)
            ec = event.get("entryConditions") or []
            today_event_class = next(
                (c["description"] for c in ec if c.get("type") == "Class"), None
            )
            for sel in event.get("selections", []):
                competitor = sel.get("competitor") or {}
                horse_name = competitor.get("name", "").strip()
                if not horse_name:
                    continue

                stats = sel.get("stats") or {}
                wins  = stats.get("wins") or 0
                total = stats.get("totalRuns") or 0
                dry   = stats.get("dryPlaces") or [0, 0, 0]
                wet   = stats.get("wetPlaces") or [0, 0, 0]

                # Normalise to lists (API sometimes returns ints for 0-count fields)
                if not isinstance(dry, list):
                    dry = [0, 0, 0]
                if not isinstance(wet, list):
                    wet = [0, 0, 0]

                entry = {
                    "career":         f"{wins}W/{total}R" if total > 0 else "",
                    "dry":            f"{dry[0]}-{dry[1]}-{dry[2]}",
                    "wet":            f"{wet[0]}-{wet[1]}-{wet[2]}",
                    "dry_arr":        dry,              # raw list for compute_wet_preference()
                    "wet_arr":        wet,              # raw list for compute_wet_preference()
                    "last_runs":      "",
                    "days_since":     None,
                    "last_distance":  None,             # filled in Phase 2
                    # Grade levelling — today's class from entryConditions,
                    # fallback to stats.class prefix (format: "3:0-1-1")
                    "current_class":  today_event_class or _parse_stats_class(stats.get("class")),
                    "last_class":     None,             # filled in Phase 2
                    # Barrier stats — list from barrierStats {name wins runs}
                    # Keyed by barrier name so build_race_summary can match today's draw
                    "barrier_stats_raw": _index_barrier_stats(stats.get("barrierStats")),
                }
                form_lookup[horse_name.upper()] = entry
                sel_by_id[sel["id"]]            = entry

    print(
        f"  Phase 1: career stats for {len(form_lookup)} runners "
        f"({len(aus_meeting_ids)} Australian meetings to enrich)"
    )

    # ── Phase 2: lastRun per Australian meeting ────────────────────────────────
    enriched = 0
    for mtg_id in aus_meeting_ids:
        try:
            lr_data = _puntapi_gql(_PUNTAPI_LASTRUN_QUERY.replace("%meeting_id%", mtg_id))
        except Exception:
            continue
        if not lr_data or "data" not in lr_data:
            continue

        lr_mtg = lr_data["data"].get("meeting") or {}
        for event in lr_mtg.get("events", []):
            for sel in event.get("selections", []):
                entry = sel_by_id.get(sel["id"])
                if entry is None:
                    comp_name = (sel.get("competitor") or {}).get("name", "").upper()
                    entry = form_lookup.get(comp_name)
                if entry is None:
                    continue

                last_run = sel.get("lastRun")
                if not last_run:
                    continue

                lr_event  = last_run.get("event") or {}
                lr_start  = lr_event.get("startTime", "")
                lr_pos    = last_run.get("finishPosition", "?")
                lr_venue  = last_run.get("meetingName", "?")
                lr_dist   = lr_event.get("distance", "")
                lr_margin = last_run.get("margin") or 0

                days_since = None
                days_label = ""
                if lr_start:
                    try:
                        lr_dt = datetime.fromisoformat(
                            lr_start.replace("Z", "+00:00")
                        ).replace(tzinfo=None)
                        days_since = (today - lr_dt).days
                        days_label = f"{days_since}d ago"
                    except Exception:
                        pass

                margin_label = f" ({lr_margin}L)" if lr_margin and lr_margin > 0 else ""
                entry["last_runs"]  = (
                    f"{lr_pos}th{margin_label} {lr_venue} {lr_dist}m ({days_label})"
                )
                entry["days_since"] = days_since

                # Store numeric distance for step-up/down calculation
                try:
                    entry["last_distance"] = (
                        int(str(lr_dist).replace("m", "").strip()) if lr_dist else None
                    )
                except (ValueError, TypeError):
                    entry["last_distance"] = None

                # Store last run class for grade levelling
                lr_ec = lr_event.get("entryConditions") or []
                entry["last_class"] = next(
                    (c["description"] for c in lr_ec if c.get("type") == "Class"), None
                )

                enriched += 1

    print(f"  Phase 2: last-run details added for {enriched} Australian runners")

    # ── Phase 3: jockey and trainer stats (experimental) ──────────────────────
    jockey_lookup  = {}
    trainer_lookup = {}
    try:
        jt_data = _puntapi_gql(_PUNTAPI_JOCKTRN_QUERY.replace("%date%", race_date))

        if jt_data and "data" in jt_data and "errors" not in jt_data:
            for meeting in jt_data["data"].get("meetings", []):
                for event in meeting.get("events", []):
                    for sel in event.get("selections", []):
                        jockey  = sel.get("jockey") or {}
                        trainer = sel.get("trainer") or {}

                        jname  = jockey.get("name", "").strip().upper()
                        jstats = jockey.get("stats") or {}
                        if jname and jstats.get("totalRuns"):
                            jockey_lookup[jname] = {
                                "wins":       jstats.get("wins", 0),
                                "total_runs": jstats["totalRuns"],
                            }

                        tname  = trainer.get("name", "").strip().upper()
                        tstats = trainer.get("stats") or {}
                        if tname and tstats.get("totalRuns"):
                            trainer_lookup[tname] = {
                                "wins":       tstats.get("wins", 0),
                                "total_runs": tstats["totalRuns"],
                            }

            if jockey_lookup or trainer_lookup:
                print(
                    f"  Phase 3: jockey stats for {len(jockey_lookup)} jockeys, "
                    f"trainer stats for {len(trainer_lookup)} trainers"
                )
            else:
                print("  Phase 3: jockey/trainer stats unavailable (PuntAPI schema may differ)")
        else:
            errs = jt_data.get("errors", []) if jt_data else []
            msg  = errs[0].get("message", "unknown")[:80] if errs else "no data"
            print(f"  Phase 3: jockey/trainer query not supported — {msg}")

    except Exception as e:
        print(f"  Phase 3: skipped ({e})")

    return form_lookup, jockey_lookup, trainer_lookup


# ============================================================
# STEP 2: GET AI PICKS FROM CLAUDE
# ============================================================

def _format_race_block(race: dict) -> str:
    """Format a single race as compact text for the AI prompt."""
    wet_flag = "  *** WET TRACK ***" if race.get("track_wet") else ""
    lines = [
        f"--- {race['track']} R{race['race_number']} | {race['race_name']} "
        f"{race['distance']}m | {race['track_condition']}{wet_flag} ---"
    ]

    for r in race["runners"]:
        # Main runner line
        weight_str = f" {r['weight']}kg" if r.get("weight") else ""
        line = (
            f"  {r['number']}. {r['name']} (B{r['barrier']}){weight_str} "
            f"J:{r['jockey']} T:{r['trainer']} "
            f"Win:${r['win_fixed']:.2f} Pl:${r['place_fixed']:.2f}"
        )

        form_parts = []

        # Career record
        if r.get("career"):
            form_parts.append(r["career"])

        # Surface preference — flag prominently when track matches preference
        wet_pref = r.get("wet_preference", "")
        if wet_pref:
            if wet_pref == "WET TRACKER" and race.get("track_wet"):
                form_parts.append("*** WET TRACKER — ADVANTAGES TODAY ***")
            elif wet_pref == "DRY PREFERRED" and race.get("track_wet"):
                form_parts.append("!! DRY PREFERRED — DISADVANTAGED TODAY")
            else:
                form_parts.append(wet_pref)

        # Dry / wet place splits
        if r.get("dry") and r["dry"] != "0-0-0":
            form_parts.append(f"Dry:{r['dry']}")
        if r.get("wet") and r["wet"] != "0-0-0":
            form_parts.append(f"Wet:{r['wet']}")

        # Days since last run — label as FRESH / RETURNING / Nd
        if r.get("days_since") is not None:
            d = r["days_since"]
            if d < 14:
                form_parts.append("FRESH")
            elif d > 60:
                form_parts.append(f"RETURNING ({d}d)")
            else:
                form_parts.append(f"{d}d")

        # Distance change vs last start
        if r.get("distance_change"):
            form_parts.append(r["distance_change"])

        # Jockey win rate
        if r.get("jockey_wins") is not None and r.get("jockey_runs"):
            pct = r["jockey_wins"] / r["jockey_runs"] * 100
            form_parts.append(
                f"J%:{r['jockey_wins']}W/{r['jockey_runs']}R({pct:.0f}%)"
            )

        # Trainer win rate
        if r.get("trainer_wins") is not None and r.get("trainer_runs"):
            pct = r["trainer_wins"] / r["trainer_runs"] * 100
            form_parts.append(
                f"T%:{r['trainer_wins']}W/{r['trainer_runs']}R({pct:.0f}%)"
            )

        # Barrier — horse's own record from today's draw
        if r.get("barrier_flag"):
            form_parts.append(r["barrier_flag"])

        # Track barrier bias note
        if r.get("track_barrier_note"):
            form_parts.append(r["track_barrier_note"])

        # Grade levelling
        if r.get("grade_change"):
            gc = r["grade_change"]
            if "DROPS" in gc:
                form_parts.append(f"*** {gc} ***")
            elif "RISES" in gc:
                form_parts.append(f"!! {gc}")
            else:
                form_parts.append(gc)

        # Racing.com data
        if r.get("speed_rating") is not None:
            form_parts.append(f"SpeedRating:{r['speed_rating']}")
        rc_track = _fmt_rc_stat(r.get("rc_track_stats"))
        if rc_track:
            form_parts.append(f"Track:{rc_track}")
        rc_dist = _fmt_rc_stat(r.get("rc_dist_stats"))
        if rc_dist:
            form_parts.append(f"Dist:{rc_dist}")
        rc_jock = _fmt_rc_stat(r.get("rc_jockey_stats"))
        if rc_jock:
            form_parts.append(f"JockeyAtVenue:{rc_jock}")
        rc_class = _fmt_rc_stat(r.get("rc_class_stats"))
        if rc_class:
            form_parts.append(f"AtClass:{rc_class}")
        if r.get("weight_change"):
            form_parts.append(r["weight_change"])

        if form_parts:
            line += f"\n    [FORM] {' | '.join(form_parts)}"

        if r.get("last_runs"):
            # Truncate to keep tokens manageable
            lr = r["last_runs"][:110]
            line += f"\n    [RUNS] {lr}"

        lines.append(line)

    return "\n".join(lines) + "\n"


def _parse_ai_response(raw: str) -> list:
    """Parse Claude's JSON response, stripping markdown fences if present."""
    raw = raw.strip()
    if raw.startswith("```"):
        parts = raw.split("```")
        raw = parts[1] if len(parts) > 1 else raw
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()
    try:
        return json.loads(raw).get("picks", [])
    except json.JSONDecodeError:
        # Attempt to recover truncated JSON
        last_close = raw.rfind("},\n    {")
        if last_close == -1:
            last_close = raw.rfind("    }")
        if last_close > 0:
            truncated = raw[:last_close + 5] + "\n  ]\n}"
            try:
                return json.loads(truncated).get("picks", [])
            except json.JSONDecodeError:
                pass
        return []


def get_ai_picks(all_races: list, race_date: str) -> dict:
    """
    Send race data to Claude in batches and get a pick + analysis for every race.
    Returns a dict keyed by "TRACK_R{number}" -> {pick, odds, rating, analysis}
    """
    client        = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    picks         = {}
    batches       = [all_races[i:i + BATCH_SIZE] for i in range(0, len(all_races), BATCH_SIZE)]
    total_batches = len(batches)

    for batch_num, batch in enumerate(batches, 1):
        race_text = f"Horse Racing — {race_date}\n\n"
        for race in batch:
            race_text += _format_race_block(race)

        is_first      = (batch_num == 1)
        best_bet_note = (
            "Mark exactly 5 of these races as ★★★ BEST BET (your strongest picks today)."
            if is_first else
            "Use ★★★ BEST BET only for outstanding value; otherwise ★★ or ★."
        )

        prompt = f"""You are an expert Australian horse racing analyst. Pick a winner for EVERY race below.

{race_text}
{best_bet_note}
Use ★★ STRONG BET for confident picks and ★ TIP for speculative picks.

Weight these factors in order of importance:
1. TRACK CONDITION — on *** WET TRACK ***, heavily favour *** WET TRACKER *** horses. Penalise !! DRY PREFERRED horses.
2. GRADE LEVELLING — *** DROPS IN CLASS *** is a strong positive signal even with ordinary recent form. !! RISES IN CLASS is a negative signal.
3. BARRIER — BARRIER ADVANTAGE = horse wins often from today's draw; BARRIER CONCERN = poor record from this draw. GOOD DRAW at inside-biased tracks (Moonee Valley, Doomben, Caulfield) is very valuable. WIDE DRAW at tight tracks is a serious disadvantage.
4. JOCKEY FORM — J% shows jockey win rate. Prefer jockeys >15% win rate; avoid <10%.
5. TRAINER FORM — T% shows trainer win rate. High-strike trainers (>20%) are strong signals.
6. DISTANCE — Step UP suits stayers; Step DOWN suits sprinters. Large steps (400m+) are risky. Dist:NW/NR(%) shows record at today's exact distance.
7. FRESHNESS — FRESH (<14d) = peak fitness. RETURNING (>60d) = fitness risk unless trainer has good fresh record.
8. CAREER RECORD — Low W/R ratio = unexposed and potentially better than odds suggest.
9. DRY/WET SPLITS — Dry:W-P-S and Wet:W-P-S show surface-specific record.
10. SPEED RATING — SpeedRating:N from Racing.com (lower = faster/better at this track/distance). Prefer SpeedRating ≤5.
11. TRACK/CLASS RECORD — Track:NW/NR(%) = record at this specific track. AtClass:NW/NR(%) = record at this class level. JockeyAtVenue:NW/NR(%) = jockey's record at this venue+distance.
12. WEIGHT — Lighter Xkg = positive (easier to carry); Heavier Xkg = negative.

Return ONLY valid JSON — no markdown fences, no other text:
{{"picks": [{{"track": "TRACK NAME", "race_number": 1, "pick": "HORSE NAME", "barrier": "N", "odds": "$X.XX", "rating": "★★★ BEST BET", "analysis": "2-3 sentences citing specific form data."}}]}}

Rules: pick a winner for EVERY race listed. No skipping."""

        print(
            f"\n  Batch {batch_num}/{total_batches} — {len(batch)} races → {CLAUDE_MODEL}..."
        )
        response = client.messages.create(
            model=CLAUDE_MODEL,
            max_tokens=8192,
            messages=[{"role": "user", "content": prompt}],
        )

        raw         = response.content[0].text
        batch_picks = _parse_ai_response(raw)

        for pick in batch_picks:
            key = f"{pick['track']}_R{pick['race_number']}"
            picks[key] = pick

        print(f"  → {len(batch_picks)} picks received (running total: {len(picks)})")

    print(f"\nTotal picks from Claude: {len(picks)}")
    return picks


# ============================================================
# STEP 3: BUILD EXCEL SPREADSHEET
# ============================================================

TRACK_COLOURS = {
    "CAULFIELD":     {"dark": "1F3864", "light": "D6E4F0", "accent": "2E75B6"},
    "ROSEHILL":      {"dark": "1F4E79", "light": "DDEEFF", "accent": "2F80C0"},
    "FLEMINGTON":    {"dark": "1F3864", "light": "D6E4F0", "accent": "2E75B6"},
    "RANDWICK":      {"dark": "1F4E79", "light": "DDEEFF", "accent": "2F80C0"},
    "DOOMBEN":       {"dark": "375623", "light": "E2EFDA", "accent": "548235"},
    "EAGLE FARM":    {"dark": "375623", "light": "E2EFDA", "accent": "548235"},
    "MORPHETTVILLE": {"dark": "4A235A", "light": "EBE0F5", "accent": "7030A0"},
    "ASCOT":         {"dark": "7B2C2C", "light": "FCE4D6", "accent": "C00000"},
    "DEFAULT":       {"dark": "243F5A", "light": "E8F0F8", "accent": "305070"},
}

# Wet track — use a blue-tinted light for wet meetings
_WET_LIGHT = "D0E8FF"


def tb():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def _wet_pref_colour(pref: str, track_wet: bool) -> str:
    """Return a fill colour for the surface preference badge cell."""
    if pref == "WET TRACKER" and track_wet:
        return "0070C0"   # bright blue — advantage
    if pref == "WET TRACKER":
        return "9DC3E6"   # soft blue
    if pref == "DRY PREFERRED" and track_wet:
        return "FF7070"   # red — disadvantage
    if pref == "DRY PREFERRED":
        return "FFE0B0"   # light amber
    return ""


def build_spreadsheet(all_races: list, picks: dict, race_date: str) -> str:
    wb = Workbook()

    # ── SUMMARY SHEET ──────────────────────────────────────────────────────────
    ws       = wb.active
    ws.title = "Summary"

    ws.merge_cells("A1:H1")
    ws["A1"] = (
        f"AUSTRALIAN RACING TIPS — {race_date.upper()}  |  "
        "AI-POWERED BY CLAUDE + RACENET FORM"
    )
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = fill("1F3864")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        "TAB live odds  |  Racenet form  |  Wet/Dry tracker  |  "
        "Jockey & trainer stats  |  AI analysis by Claude"
    )
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 15

    headers = ["Track", "State", "Track Cond.", "Races", "Best Bet", "Odds", "Rating"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill      = fill("2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = tb()
    ws.row_dimensions[4].height = 18

    tracks_seen = {}
    for race in all_races:
        t = race["track"]
        if t not in tracks_seen:
            tracks_seen[t] = {
                "location":        race["location"],
                "track_condition": race["track_condition"],
                "track_wet":       race["track_wet"],
                "count":           0,
                "best":            None,
                "best_odds":       None,
                "best_rating":     None,
            }
        tracks_seen[t]["count"] += 1
        key = f"{t}_R{race['race_number']}"
        if key in picks and "★★★" in picks[key].get("rating", ""):
            if tracks_seen[t]["best"] is None:
                tracks_seen[t]["best"]       = picks[key]["pick"]
                tracks_seen[t]["best_odds"]  = picks[key]["odds"]
                tracks_seen[t]["best_rating"] = picks[key]["rating"]

    for i, (track, info) in enumerate(tracks_seen.items()):
        r  = i + 5
        bg = "EAF2FF" if info["track_wet"] else ("F2F7FF" if i % 2 == 0 else "FFFFFF")
        row_data = [
            track,
            info["location"],
            info["track_condition"],
            info["count"],
            info["best"] or "—",
            info["best_odds"] or "—",
            info["best_rating"] or "★",
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(
                name="Arial", bold=(col == 5), size=10,
                color="C00000" if col == 7 and "★★★" in str(val) else "000000",
            )
            c.fill      = fill(bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = tb()
        ws.row_dimensions[r].height = 17

    for i, w in enumerate([20, 8, 14, 7, 24, 10, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # ── ONE SHEET PER TRACK ────────────────────────────────────────────────────
    for track_name, tinfo in tracks_seen.items():
        colours = TRACK_COLOURS.get(track_name, TRACK_COLOURS["DEFAULT"])
        dark, light, accent = colours["dark"], colours["light"], colours["accent"]
        if tinfo["track_wet"]:
            light = _WET_LIGHT   # blue tint for wet meetings

        ws = wb.create_sheet(track_name[:24])
        ws.merge_cells("A1:J1")
        ws["A1"] = (
            f"  {track_name}  —  {race_date}  —  "
            f"{tinfo['track_condition']}{'  [WET TRACK]' if tinfo['track_wet'] else ''}"
            "  —  AI TIPS"
        )
        ws["A1"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws["A1"].fill      = fill(dark)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26

        row         = 2
        track_races = [r for r in all_races if r["track"] == track_name]

        for race in track_races:
            key          = f"{track_name}_R{race['race_number']}"
            pick_data    = picks.get(key, {})
            pick_name    = pick_data.get("pick", "No pick")
            pick_odds    = pick_data.get("odds", "—")
            pick_rating  = pick_data.get("rating", "★")
            pick_analysis = pick_data.get("analysis", "Market selection.")

            # Race header row
            ws.merge_cells(f"A{row}:J{row}")
            ws[f"A{row}"] = (
                f"  RACE {race['race_number']}  ·  {race['race_name']}  "
                f"{race['distance']}m  ·  PICK: {pick_name}  {pick_odds}  {pick_rating}"
            )
            ws[f"A{row}"].font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            ws[f"A{row}"].fill      = fill(accent)
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 19
            row += 1

            # Column headers (10 columns: A-J)
            col_heads = [
                "No.", "Horse", "Barrier", "Jockey",
                "Win", "Place", "Surface", "Career/Stats", "Grade", "Last Run",
            ]
            for col, h in enumerate(col_heads, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font      = Font(name="Arial", bold=True, size=8, color="FFFFFF")
                c.fill      = fill(dark)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = tb()
            ws.row_dimensions[row].height = 13
            row += 1

            # Runner rows
            for j, runner in enumerate(race["runners"]):
                is_pick  = (runner["name"] == pick_name)
                rbg      = "FFF2CC" if is_pick else ("F9F9F9" if j % 2 == 0 else "FFFFFF")
                wet_pref = runner.get("wet_preference", "")

                # Build career/stats cell text (col 8)
                stats_parts = []
                if runner.get("career"):
                    stats_parts.append(runner["career"])
                if runner.get("distance_change"):
                    stats_parts.append(runner["distance_change"])
                jw, jr = runner.get("jockey_wins"), runner.get("jockey_runs")
                if jw is not None and jr:
                    stats_parts.append(f"J:{jw}W/{jr}R({jw/jr*100:.0f}%)")
                tw, tr_ = runner.get("trainer_wins"), runner.get("trainer_runs")
                if tw is not None and tr_:
                    stats_parts.append(f"T:{tw}W/{tr_}R({tw/tr_*100:.0f}%)")
                # Barrier stats in career/stats cell
                if runner.get("barrier_flag"):
                    stats_parts.append(runner["barrier_flag"])
                elif runner.get("track_barrier_note"):
                    stats_parts.append(runner["track_barrier_note"])
                # Racing.com stats
                if runner.get("speed_rating") is not None:
                    stats_parts.append(f"Speed:{runner['speed_rating']}")
                rc_t = _fmt_rc_stat(runner.get("rc_track_stats"))
                rc_d = _fmt_rc_stat(runner.get("rc_dist_stats"))
                if rc_t:
                    stats_parts.append(f"Track:{rc_t}")
                if rc_d:
                    stats_parts.append(f"Dist:{rc_d}")
                if runner.get("weight_change"):
                    stats_parts.append(runner["weight_change"])
                stats_text = "\n".join(stats_parts) if stats_parts else ""

                # Grade change cell (col 9)
                grade_text = runner.get("grade_change", "")

                vals = [
                    runner["number"],           # 1  No.
                    runner["name"],             # 2  Horse
                    runner["barrier"],          # 3  Barrier
                    runner["jockey"],           # 4  Jockey
                    f"${runner['win_fixed']:.2f}" if runner["win_fixed"] else "—",   # 5 Win
                    f"${runner['place_fixed']:.2f}" if runner["place_fixed"] else "—",  # 6 Place
                    wet_pref,                   # 7  Surface
                    stats_text,                 # 8  Career/Stats
                    grade_text,                 # 9  Grade
                    runner.get("last_runs", ""),  # 10 Last Run
                ]

                for col, val in enumerate(vals, 1):
                    c        = ws.cell(row=row, column=col, value=val)
                    c.border = tb()
                    c.alignment = Alignment(
                        horizontal="left" if col in [2, 4, 7, 8, 9, 10] else "center",
                        vertical="center",
                        wrap_text=(col in [8, 9, 10]),
                    )

                    # Surface preference cell — coloured background
                    if col == 7 and wet_pref:
                        pref_fill = _wet_pref_colour(wet_pref, race["track_wet"])
                        c.fill = fill(pref_fill) if pref_fill else fill(rbg)
                        c.font = Font(
                            name="Arial", bold=True, size=8,
                            color="FFFFFF" if pref_fill in ("0070C0", "FF7070") else "000000",
                        )
                    # Grade change cell — colour-coded
                    elif col == 9 and grade_text:
                        if "DROPS" in grade_text:
                            c.fill = fill("C6EFCE")   # green — positive
                            c.font = Font(name="Arial", bold=True, size=9, color="276221")
                        elif "RISES" in grade_text:
                            c.fill = fill("FFCCCC")   # red — negative
                            c.font = Font(name="Arial", bold=True, size=9, color="9C0006")
                        else:
                            c.fill = fill(rbg)
                            c.font = Font(name="Arial", size=9, color="666666")
                    else:
                        c.fill = fill(rbg)
                        c.font = Font(
                            name="Arial", bold=is_pick,
                            size=9 if col in [8, 9, 10] else 10,
                            color="C00000" if (is_pick and col == 2) else "000000",
                        )

                has_wrap = bool(runner.get("last_runs")) or bool(stats_text) or bool(grade_text)
                ws.row_dimensions[row].height = 35 if has_wrap else 17
                row += 1

            # AI analysis row
            ws.merge_cells(f"A{row}:J{row}")
            ws[f"A{row}"] = f"  Analysis: {pick_analysis}"
            ws[f"A{row}"].font      = Font(name="Arial", italic=True, size=9, color="404040")
            ws[f"A{row}"].fill      = fill(light)
            ws[f"A{row}"].alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
            ws.row_dimensions[row].height = 15
            row += 1

            # Result placeholder
            ws.merge_cells(f"A{row}:J{row}")
            ws[f"A{row}"] = "  RESULT: ___________________________________________"
            ws[f"A{row}"].font      = Font(name="Arial", size=9, color="AAAAAA")
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 13
            row += 1
            ws.row_dimensions[row].height = 5
            row += 1

        # Column widths: No, Horse, Barrier, Jockey, Win, Place, Surface, Career/Stats, Grade, Last Run
        for col, w in enumerate([5, 22, 7, 18, 9, 9, 14, 22, 18, 34], 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A2"

    filename = f"TAB_Racing_Tips_{race_date}.xlsx"
    wb.save(filename)
    return filename


# ============================================================
# STEP 4: EMAIL DELIVERY
# ============================================================

def send_email(filename: str, race_date: str) -> bool:
    """
    Send the Excel file as an email attachment via Gmail SMTP.
    Requires GMAIL_USER, GMAIL_APP_PASSWORD, and EMAIL_RECIPIENTS env vars.
    Returns True on success, False if unconfigured or failed.

    Gmail setup:
      1. Enable 2-Step Verification on your Google account
      2. Go to Google Account > Security > App Passwords
      3. Create an App Password for "Mail / Windows Computer"
      4. Use that 16-character password as GMAIL_APP_PASSWORD
    """
    if not GMAIL_USER or not GMAIL_APP_PASS or not EMAIL_RECIPIENTS:
        print(
            "\n  Email not configured — set GMAIL_USER, GMAIL_APP_PASSWORD, "
            "EMAIL_RECIPIENTS environment variables to enable."
        )
        return False

    print(f"\n  Sending email to: {', '.join(EMAIL_RECIPIENTS)}...")

    msg             = MIMEMultipart()
    msg["From"]     = GMAIL_USER
    msg["To"]       = ", ".join(EMAIL_RECIPIENTS)
    msg["Subject"]  = f"TAB Racing Tips — {race_date}"

    body = (
        f"Racing tips for {race_date} are attached.\n\n"
        "Generated by TAB Racing Bot — AI analysis by Claude + Racenet form data.\n"
    )
    msg.attach(MIMEText(body, "plain"))

    try:
        with open(filename, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{os.path.basename(filename)}"',
        )
        msg.attach(part)
    except OSError as e:
        print(f"  Email failed — could not read file: {e}")
        return False

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(GMAIL_USER, GMAIL_APP_PASS)
            server.sendmail(GMAIL_USER, EMAIL_RECIPIENTS, msg.as_string())
        print(f"  Email sent successfully to {len(EMAIL_RECIPIENTS)} recipient(s).")
        return True
    except smtplib.SMTPAuthenticationError:
        print(
            "  Email failed — authentication error. "
            "Check GMAIL_USER and GMAIL_APP_PASSWORD (use an App Password, not your login)."
        )
        return False
    except Exception as e:
        print(f"  Email failed: {e}")
        return False


# ============================================================
# STEP 5: WINDOWS TASK SCHEDULER SETUP
# ============================================================

def setup_task_scheduler():
    """
    Create a batch file for Windows Task Scheduler and print setup instructions.
    Run with:  python tab_racing_bot.py --setup
    """
    script_dir  = os.path.dirname(os.path.abspath(__file__))
    script_path = os.path.abspath(__file__)
    bat_path    = os.path.join(script_dir, "run_racing_bot.bat")

    bat_content = (
        f"@echo off\r\n"
        f"cd /d \"{script_dir}\"\r\n"
        f"echo Running TAB Racing Bot...\r\n"
        f"python \"{script_path}\"\r\n"
        f"echo Done. Check for TAB_Racing_Tips_*.xlsx in {script_dir}\r\n"
        f"pause\r\n"
    )

    with open(bat_path, "w", encoding="utf-8") as f:
        f.write(bat_content)

    print(f"\nBatch file created: {bat_path}")
    print("\n" + "=" * 60)
    print("  WINDOWS TASK SCHEDULER SETUP")
    print("=" * 60)
    print("\nTo run every Saturday at 7:00 AM automatically:")
    print("\n  1. Press Win+S, search 'Task Scheduler', open it")
    print("  2. Click 'Create Basic Task...' in the right panel")
    print("  3. Name: TAB Racing Bot   Description: Weekly racing tips")
    print("  4. Click Next → select 'Weekly' → Next")
    print("  5. Set Start time: 7:00 AM")
    print("  6. Check 'Saturday' only → Next")
    print("  7. Select 'Start a program' → Next")
    print(f"  8. Program/script: {bat_path}")
    print("  9. Click Finish")
    print("\nIMPORTANT: Make sure ANTHROPIC_API_KEY is set as a system")
    print("environment variable (not just user), or the task won't find it.")
    print("To set a system variable:")
    print("  1. Win+S → 'Edit the system environment variables'")
    print("  2. Click 'Environment Variables' → under System Variables → New")
    print("  3. Name: ANTHROPIC_API_KEY  Value: sk-ant-api03-...")
    print("\n" + "=" * 60)


# ============================================================
# MAIN
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="TAB Racing Bot — AI-powered tips with Racenet form"
    )
    parser.add_argument(
        "--date", default=str(date.today()),
        help="Race date YYYY-MM-DD (default: today)",
    )
    parser.add_argument(
        "--state", default="NSW",
        help="TAB jurisdiction (NSW = all national meetings)",
    )
    parser.add_argument(
        "--setup", action="store_true",
        help="Create Windows Task Scheduler batch file and print setup instructions",
    )
    parser.add_argument(
        "--no-email", action="store_true",
        help="Skip sending email even if email env vars are configured",
    )
    parser.add_argument(
        "--all-tracks", action="store_true",
        help="Disable country whitelist and fetch all international meetings",
    )
    args = parser.parse_args()

    if args.setup:
        setup_task_scheduler()
        return

    all_tracks = getattr(args, "all_tracks", False)
    tracks_label = "ALL COUNTRIES" if all_tracks else "AU + NZ + JPN"

    print(f"\n{'='*60}")
    print(f"  TAB RACING BOT  |  Date: {args.date}  |  State: {args.state}")
    print(f"  Model: {CLAUDE_MODEL}  |  Tracks: {tracks_label}")
    print(f"{'='*60}\n")

    # 1. Fetch TAB meetings (whitelist applied unless --all-tracks)
    meetings = get_meetings(args.date, args.state, all_tracks=all_tracks)
    if not meetings:
        print("No meetings found. Check the date or jurisdiction.")
        return

    # 2. Fetch Racenet form, jockey, and trainer data
    form_lookup, jockey_lookup, trainer_lookup = get_puntapi_form(args.date)

    # 2b. Fetch Racing.com speed ratings and form breakdowns
    racing_com_lookup = get_racing_com_form(args.date)

    # 3. Fetch all TAB race details and enrich with form data
    all_races = build_race_summary(
        meetings, args.date, args.state,
        form_lookup, jockey_lookup, trainer_lookup,
        racing_com_lookup,
    )

    enriched_form    = sum(1 for r in all_races for runner in r["runners"] if runner.get("career"))
    enriched_wet     = sum(1 for r in all_races for runner in r["runners"] if runner.get("wet_preference"))
    enriched_jockey  = sum(1 for r in all_races for runner in r["runners"] if runner.get("jockey_wins") is not None)
    enriched_trainer = sum(1 for r in all_races for runner in r["runners"] if runner.get("trainer_wins") is not None)
    enriched_barrier = sum(1 for r in all_races for runner in r["runners"] if runner.get("barrier_wins") is not None)
    enriched_grade   = sum(1 for r in all_races for runner in r["runners"] if runner.get("grade_change"))
    enriched_bias    = sum(1 for r in all_races for runner in r["runners"] if runner.get("track_barrier_note"))
    enriched_rc      = sum(1 for r in all_races for runner in r["runners"] if runner.get("speed_rating") is not None)
    enriched_wt      = sum(1 for r in all_races for runner in r["runners"] if runner.get("weight_change"))
    wet_meetings     = sum(1 for r in all_races if r["track_wet"])

    print(f"\nTotal races fetched:          {len(all_races)}")
    print(f"Wet track meetings:           {wet_meetings}")
    print(f"Runners with career form:     {enriched_form}")
    print(f"Runners with surface pref:    {enriched_wet}")
    print(f"Runners with jockey stats:    {enriched_jockey}")
    print(f"Runners with trainer stats:   {enriched_trainer}")
    print(f"Runners with barrier stats:   {enriched_barrier}")
    print(f"Runners with grade change:    {enriched_grade}")
    print(f"Runners with track bias note: {enriched_bias}")
    print(f"Runners with Racing.com data: {enriched_rc}")
    print(f"Runners with weight change:   {enriched_wt}")

    # 4. Get AI picks from Claude
    picks = get_ai_picks(all_races, args.date)

    # 5. Build Excel spreadsheet
    filename = build_spreadsheet(all_races, picks, args.date)
    print(f"\nDone! Spreadsheet saved: {filename}")
    print(f"    Races covered: {len(all_races)}")
    print(f"    Picks made:    {len(picks)}")

    # 6. Send email (if configured and not suppressed)
    if not args.no_email:
        send_email(filename, args.date)
    else:
        print("\n  --no-email flag set, skipping email.")


if __name__ == "__main__":
    main()
